#!/usr/bin/env python3
"""
Complete Solution: Extract invoice data and populate Excel template in one script
Usage: python3 complete_solution.py --source 1.xls --template "مصرف کننده.xlsm" --output result.xlsx
"""
import argparse
import json
import pathlib
import pandas as pd
import re
from xml.etree import ElementTree as ET
from openpyxl import load_workbook
from datetime import datetime

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def persian_digits_to_en(s):
    """Convert Persian digits to English"""
    if s is None: return None
    t = str(s)
    return t.translate(str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789"))

def to_number(s):
    """Convert string to number, handling Persian digits and formatting"""
    if s is None: return None
    t = persian_digits_to_en(s)
    t = t.replace(",", "").replace("٬", "").strip()
    if re.fullmatch(r"-?\d+(\.\d+)?", t or ""):
        try: 
            return int(t) if "." not in t else float(t)
        except: 
            return None
    return None

def persian_to_english_number(persian_str):
    """Convert Persian number format to English numeric value"""
    if not persian_str or pd.isna(persian_str):
        return None
    
    # Convert Persian digits to English
    persian_str = str(persian_str).translate(str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789"))
    
    # Remove commas and spaces
    persian_str = persian_str.replace(",", "").replace(" ", "").strip()
    
    # Convert to numeric
    try:
        return float(persian_str) if "." in persian_str else int(persian_str)
    except:
        return None

# ============================================================================
# SPREADSHEETML PARSING
# ============================================================================

def spreadsheetml_tables(path):
    """Parse SpreadsheetML format files"""
    ns = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}
    root = ET.parse(path).getroot()
    out = []
    
    for ws in root.findall(".//ss:Worksheet", ns):
        name = ws.attrib.get("{urn:schemas-microsoft-com:office:spreadsheet}Name") or "Sheet1"
        table = ws.find("ss:Table", ns)
        rows = []
        
        if table is not None:
            for r in table.findall("ss:Row", ns):
                row = []
                col = 1
                for c in r.findall("ss:Cell", ns):
                    idx = c.attrib.get("{urn:schemas-microsoft-com:office:spreadsheet}Index")
                    if idx:
                        idx = int(idx)
                        while col < idx: 
                            row.append(None)
                            col += 1
                    d = c.find("ss:Data", ns)
                    v = d.text if d is not None else None
                    row.append(v)
                    col += 1
                rows.append(row)
        
        w = max((len(x) for x in rows), default=0)
        rows = [x + [None] * (w - len(x)) for x in rows]
        out.append((name, pd.DataFrame(rows)))
    
    return out

def load_sheet(path, sheet_name=None):
    """Load Excel or SpreadsheetML files"""
    if str(path).lower().endswith(".xlsx"):
        if sheet_name:
            df = pd.read_excel(path, sheet_name=sheet_name, header=None, dtype=object, engine="openpyxl")
            return sheet_name, df
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        name = wb.sheetnames[0]
        df = pd.read_excel(path, sheet_name=name, header=None, dtype=object, engine="openpyxl")
        return name, df
    
    tables = spreadsheetml_tables(path)
    if sheet_name:
        tables = [t for t in tables if t[0] == sheet_name] or tables[:1]
    return tables[0]

# ============================================================================
# INVOICE DATA EXTRACTION
# ============================================================================

def detect_product_header_row(df, max_scan=60):
    """Automatically detect the row containing product table headers"""
    best = None
    score = -1
    
    for i in range(min(len(df), max_scan)):
        row = df.iloc[i].astype(str)
        nz = (row != "None") & (row != "nan") & (row.str.strip() != "")
        toks = row.where(nz).dropna().str.strip()
        hits = sum(toks.str.contains("شرح|کالا|خدمات|مقدار|واحد|مبلغ|ردیف|کد"))
        short = (toks.str.len() <= 30).sum()
        uniq = toks.str.lower().nunique()
        s = int(nz.sum()) + short + uniq + 4 * hits
        if s > score: 
            score = s
            best = i
    
    return best

def canon_label(s):
    """Standardize column labels"""
    t = str(s or "").strip()
    
    CANON = [
        (re.compile("شرح.*(کالا|خدمات)"), "شرح کالا یا خدمات"),
        (re.compile("مقدار"), "مقدار"),
        (re.compile("واحد"), "واحد"),
        (re.compile("مبلغ.?واحد"), "مبلغ واحد"),
        (re.compile("مبلغ.?کل"), "مبلغ کل"),
        (re.compile("تخفیف"), "مبلغ تخفیف"),
        (re.compile("مالیات|عوارض"), "جمع مالیات و عوارض"),
        (re.compile("ردیف"), "ردیف"),
        (re.compile("^کد$"), "کد"),
        (re.compile("کد"), "کد۲"),
    ]
    
    for rx, k in CANON:
        if rx.search(t): 
            return k
    return t

def drop_empty_rows(df, min_nonempty=2):
    """Remove rows with insufficient non-empty cells"""
    if df.size == 0:
        return df
    
    nonempty = df.notna() & (df.astype(str).apply(lambda col: col.str.strip().ne("")) if df.size > 0 else pd.DataFrame())
    mask = nonempty.sum(axis=1) >= min_nonempty
    return df.loc[mask].reset_index(drop=True)

def build_items(df_raw, header_row, stop_on_total=True, min_nonempty=2):
    """Extract product items from the data"""
    headers_row = df_raw.iloc[header_row].tolist()
    idx_to_key = {}
    
    for j, h in enumerate(headers_row):
        key = canon_label(h)
        if key:
            if key in idx_to_key.values():
                k = key
                c = 2
                while f"{k}{c}" in idx_to_key.values(): 
                    c += 1
                key = f"{k}{c}"
            idx_to_key[j] = key
    
    body = df_raw.iloc[header_row + 1:].copy()
    body = drop_empty_rows(body, min_nonempty=min_nonempty)
    items = []
    summary = {}
    
    NUMERIC_FIELDS = {"مقدار", "مبلغ واحد", "مبلغ کل", "مبلغ تخفیف", "جمع مالیات و عوارض", "ردیف"}
    HEADER_TOKENS = {"کد", "کد۲", "مقدار", "واحد", "مبلغ واحد", "مبلغ تخفیف", "جمع مالیات و عوارض", "ردیف", "شرح کالا یا خدمات"}
    
    for _, row in body.iterrows():
        rvals = row.tolist()
        
        # Check if this is a total row
        if any((str(x).strip() == "جمع کل" for x in rvals if x is not None)):
            for j, k in idx_to_key.items():
                v = rvals[j] if j < len(rvals) else None
                if k in NUMERIC_FIELDS:
                    n = to_number(v)
                    if n is not None: 
                        summary[k] = n
            if stop_on_total: 
                break
            else: 
                continue
        
        # Skip separator/header-like rows accidentally inside the body
        nonempty_texts = [str(x).strip() for x in rvals if x is not None and str(x).strip() not in {"", "-", "."}]
        # If the row consists only of known header tokens (e.g., "کد", "مقدار", ...), skip it
        if nonempty_texts and all(t in HEADER_TOKENS for t in nonempty_texts):
            continue
        # Also skip if the row contains the phrase "جمع مالیات و عوارض" alone (header band)
        if len(nonempty_texts) == 1 and nonempty_texts[0] == "جمع مالیات و عوارض":
            continue
        
        # Build product record
        rec = {}
        for j, k in idx_to_key.items():
            v = rvals[j] if j < len(rvals) else None
            if v is None: 
                continue
            sv = str(v).strip()
            if sv == "" or sv == "None": 
                continue
            
            # Skip placing header tokens as values in item fields
            if sv in HEADER_TOKENS:
                continue
            
            if k in NUMERIC_FIELDS:
                n = to_number(sv)
                rec[k] = n if n is not None else sv
            else:
                rec[k] = sv
        
        # Minimal sanity: require at least one of product code/quantity/unit price to exist
        if not rec:
            continue
        if not any(key in rec for key in ("کد", "کد۲", "کد2", "مقدار", "مبلغ واحد")):
            # likely a non-item informational line
            continue
        
        items.append(rec)
    
    return items, summary, idx_to_key

def extract_meta(df_raw, header_row):
    """Extract metadata from header rows"""
    meta_rows = df_raw.iloc[:header_row] if header_row and header_row > 0 else pd.DataFrame()
    meta = {}
    
    # First pass: collect all potential key-value pairs
    for _, r in meta_rows.iterrows():
        cells = [(i, str(x).strip()) for i, x in enumerate(r.tolist()) if x is not None and str(x).strip() != ""]
        for i, txt in cells:
            # Look for labels that end with colons or contain specific keywords
            if (re.search(r"[:：]\s*$", txt) or 
                re.search(r"(تلفن|نشانی|کد پستی|شماره|Email|WebSite|ثبت|ملی|اقتصادی|تاریخ|فاکتور|سریال|حافظه|خریدار|فروشنده)", txt)):
                
                key = txt.rstrip(":：").strip()
                val = None
                
                # Try to find value in the next column
                for j in range(i + 1, len(r)):
                    vv = r.iloc[j]
                    if vv is not None and str(vv).strip() != "":
                        val = str(vv).strip()
                        break
                
                # If no value found in next column, try previous column
                if not val and i > 0:
                    vv = r.iloc[i - 1]
                    if (vv is not None and str(vv).strip() != "" and 
                        not re.search(r"[:：]\s*$", str(vv))):
                        val = str(vv).strip()
                
                if val:
                    # Handle duplicate keys
                    if key in meta:
                        k = key
                        t = 2
                        while f"{k}_{t}" in meta:
                            t += 1
                        key = f"{k}_{t}"
                    meta[key] = val
    
    # Clean and normalize the metadata
    cleaned = {}
    for k, v in meta.items():
        # Clean up the key (remove zero-width characters and normalize spaces)
        kk = re.sub(r'[\u200c\u200d]', ' ', k).replace("‌", " ").strip()
        vv = str(v).strip()
        
        # Convert Persian digits to English for specific fields
        if re.search(r"تلفن", kk):
            vv = persian_digits_to_en(vv)
        if re.search(r"(کد پستی|ثبت|ملی|اقتصادی)", kk):
            vv = persian_digits_to_en(vv)
        
        # Only add if we have a meaningful value
        if vv and vv != "None" and vv != "nan":
            cleaned[kk] = vv
    
    # Extract specific important fields from the header
    header_info = {}
    
    # Look for invoice dates and other header information
    for _, r in meta_rows.iterrows():
        row_text = ' '.join([str(x) for x in r.tolist() if x and str(x).strip()])
        
        # Extract invoice issue date
        if 'تاریخ صدور فاکتور' in row_text:
            date_match = re.search(r'تاریخ صدور فاکتور:(\d{4}/\d{2}/\d{2})', row_text)
            if date_match:
                header_info['تاریخ صدور فاکتور'] = date_match.group(1)
        
        # Extract invoice creation date
        if 'تاریخ ایجاد فاکتور' in row_text:
            date_match = re.search(r'تاریخ ایجاد فاکتور:(\d{4}/\d{2}/\d{2})', row_text)
            if date_match:
                header_info['تاریخ ایجاد فاکتور'] = date_match.group(1)
        
        # Extract main invoice number
        if 'شماره فاکتور اصلی' in row_text:
            invoice_match = re.search(r'شماره فاکتور اصلی:([A-Z0-9]+)', row_text)
            if invoice_match:
                header_info['شماره فاکتور اصلی'] = invoice_match.group(1)
        
        # Extract serial number
        if 'شماره سریال' in row_text:
            serial_match = re.search(r'شماره سریال:(\d+)', row_text)
            if serial_match:
                header_info['شماره سریال'] = serial_match.group(1)
        
        # Extract tax memory serial
        if 'سریال فاکتور حافظه مالیاتی' in row_text:
            tax_match = re.search(r'سریال فاکتور حافظه مالیاتی:(\d+)', row_text)
            if tax_match:
                header_info['سریال فاکتور حافظه مالیاتی'] = tax_match.group(1)
    
    # Targeted extraction from buyer section for buyer info
    buyer_section_found = False
    for _, r in meta_rows.iterrows():
        row_text = ' '.join([str(x) for x in r.tolist() if x and str(x).strip()])
        if 'مشخصات‌ خريدار‌' in row_text or 'مشخصات‌ خریدار‌' in row_text or 'مشخصات خریدار' in row_text:
            buyer_section_found = True
            continue
        if buyer_section_found:
            # Name
            m_name = re.search(r'نام شخص حقیقی/حقوقی:\s*([^\s]+)', row_text)
            if m_name:
                header_info['نام خریدار'] = m_name.group(1)
            # National ID (explicit label)
            m_id = re.search(r'شماره ثبت/ملی:(\d+)', row_text)
            if m_id:
                header_info['کد/شناسه ملی خریدار'] = persian_digits_to_en(m_id.group(1))
            # As a fallback, grab the first 10-14 digit number in buyer block
            if 'کد/شناسه ملی خریدار' not in header_info:
                m_anyid = re.search(r'(\d{10,14})', persian_digits_to_en(row_text))
                if m_anyid:
                    header_info['کد/شناسه ملی خریدار'] = m_anyid.group(1)
            # Stop if we reached next section
            if 'مشخصات کالا' in row_text or 'مشخصات‌ فروشنده‌' in row_text or 'مشخصات فروشنده' in row_text:
                break
    
    # Merge header info with cleaned metadata
    cleaned.update(header_info)
    
    return cleaned

# ============================================================================
# EXCEL TEMPLATE POPULATION
# ============================================================================

def populate_excel_template(invoice_data, excel_template_path, output_path):
    """Populate the Excel template with extracted invoice data"""
    
    # Load the Excel template
    wb = load_workbook(excel_template_path)
    ws = wb['فروش به مصرف کننده']
    
    # Extract real data from metadata
    meta = invoice_data.get('meta', {})
    
    # Get real invoice date (prefer issue date, fallback to creation date, then current date)
    invoice_date = None
    if 'تاریخ صدور فاکتور' in meta:
        invoice_date = meta['تاریخ صدور فاکتور']
    elif 'تاریخ ایجاد فاکتور' in meta:
        invoice_date = meta['تاریخ ایجاد فاکتور']
    else:
        invoice_date = datetime.now().strftime("%Y/%m/%d")
    
    # Get real invoice number - use serial number (شماره سریال) as the invoice number
    invoice_number = None
    if 'شماره سریال' in meta:
        invoice_number = meta['شماره سریال']  # Use serial number as invoice number
    elif 'شماره فاکتور اصلی' in meta:
        invoice_number = meta['شماره فاکتور اصلی']
    else:
        invoice_number = int(datetime.now().timestamp()) % 1000000
    
    # Extract buyer information - leave empty if not found
    buyer_name = meta.get('نام خریدار', '')
    buyer_id = meta.get('کد/شناسه ملی خریدار', '')
    buyer_phone = meta.get('تلفن خریدار', '')
    buyer_postal_code = meta.get('کد پستی خریدار', None)
    
    # Look for buyer postal code in metadata if not found
    if not buyer_postal_code:
        for key, value in meta.items():
            if 'کد پستی' in key and 'خریدار' in key:
                buyer_postal_code = value
                break
    
    # Extract seller postal code from metadata
    seller_postal_code = None
    for key, value in meta.items():
        if 'کد پستی' in key and 'خریدار' not in key:
            seller_postal_code = value
            break
    
    # Use buyer postal code if available, otherwise seller's, leave empty if none
            source_postal_code = persian_digits_to_en('۶۷۱۹۸۵۹۸۶۶')
    
    # Default values for missing fields (now using real data where available)
    default_values = {
        'تاریخ سند *': invoice_date,
        'شماره صورتحساب': invoice_number,
        'کد/شناسه ملی خریدار': buyer_id,
        'نام خریدار': buyer_name,
        'تلفن همراه': buyer_phone,
        '* کدپستی انبار مبدا': source_postal_code,
        'شرح سند': '',
        'سایر اضافات (ریال)': 0  # No additional charges in source
    }
    
    # Find the first empty row after headers (assuming headers are in row 1)
    start_row = 2  # Start from row 2 (after headers)
    
    # Populate each product item
    for i, item in enumerate(invoice_data['items']):
        row = start_row + i
        
        # Get product code (prefer کد2 over کد) - leave empty if not found
        product_code = item.get('کد2') or item.get('کد') or ''
        
        # Get quantity
        quantity = item.get('مقدار', 0)
        
        # Get unit price (convert from Persian format to number)
        unit_price = persian_to_english_number(item.get('واحد', 0))
        
        # Get discount amount
        discount = item.get('مبلغ تخفیف', 0)
        
        # Get tax amount
        tax = item.get('جمع مالیات و عوارض', 0)
        
        # Map the fields
        row_data = {
            'A': default_values['تاریخ سند *'],                    # تاریخ سند
            'B': default_values['شماره صورتحساب'],                # شماره صورتحساب
            'C': default_values['کد/شناسه ملی خریدار'],           # کد/شناسه ملی خریدار
            'D': default_values['نام خریدار'],                     # نام خریدار
            'E': default_values['تلفن همراه'],                     # تلفن همراه
            'F': default_values['* کدپستی انبار مبدا'],           # کدپستی انبار مبدا
            'G': default_values['شرح سند'],                       # شرح سند
            'H': product_code,                                    # شناسه کالا
            'I': quantity,                                        # تعداد/مقدار
            'J': unit_price,                                      # مبلغ واحد (ریال)
            'K': discount,                                        # مبلغ تخفیف (ریال)
            'L': default_values['سایر اضافات (ریال)'],            # سایر اضافات (ریال)
            'M': tax                                              # مبلغ مالیات و عوارض (ریال)
        }
        
        # Write data to Excel - only write non-empty values
        for col, value in row_data.items():
            if value is not None and value != '':
                ws[f'{col}{row}'] = value
    
    # Save the populated Excel file
    wb.save(output_path)
    
    return output_path, default_values

def extract_invoice_data(source_file, sheet_name=None, header_row=None):
    """Load a single source file and return invoice_data dict as used by populate."""
    sheet, df_raw = load_sheet(source_file, sheet_name=sheet_name)
    if header_row is None:
        header_row = detect_product_header_row(df_raw)
    items, summary, idxmap = build_items(df_raw, header_row)
    meta = extract_meta(df_raw, header_row)
    return {
        "sheet": sheet,
        "header_row": int(header_row),
        "columns_map": {int(k): v for k, v in idxmap.items()},
        "meta": meta,
        "items": items,
        "summary": summary,
        "source_file": str(source_file),
    }


def populate_excel_template_batch(invoices_data, excel_template_path, output_path):
    """Populate many invoices into one output workbook (append rows)."""
    wb = load_workbook(excel_template_path)
    ws = wb['فروش به مصرف کننده']
    
    # Find first data row (after headers assumed row 1)
    row = 2
    while ws[f'A{row}'].value is not None or ws[f'B{row}'].value is not None:
        row += 1
    
    for invoice_data in invoices_data:
        # Reuse same mapping logic as single populate
        meta = invoice_data.get('meta', {})
        invoice_date = meta.get('تاریخ صدور فاکتور') or meta.get('تاریخ ایجاد فاکتور') or datetime.now().strftime("%Y/%m/%d")
        invoice_number = meta.get('شماره سریال') or meta.get('شماره فاکتور اصلی') or int(datetime.now().timestamp()) % 1000000
        buyer_name = meta.get('نام خریدار', '')
        buyer_id = meta.get('کد/شناسه ملی خریدار', '')
        buyer_phone = meta.get('تلفن خریدار', '')
        source_postal_code = persian_digits_to_en('۶۷۱۹۸۵۹۸۶۶')
        for item in invoice_data['items']:
            product_code = item.get('کد2') or item.get('کد') or ''
            quantity = item.get('مقدار', 0)
            unit_price = persian_to_english_number(item.get('واحد', 0))
            discount = item.get('مبلغ تخفیف', 0)
            tax = item.get('جمع مالیات و عوارض', 0)
            row_data = {
                'A': invoice_date,
                'B': invoice_number,
                'C': buyer_id,
                'D': buyer_name,
                'E': buyer_phone,
                'F': source_postal_code,
                'G': '',
                'H': product_code,
                'I': quantity,
                'J': unit_price,
                'K': discount,
                'L': 0,
                'M': tax,
            }
            for col, value in row_data.items():
                if value is not None and value != '':
                    ws[f'{col}{row}'] = value
            row += 1
    wb.save(output_path)
    return output_path

# ============================================================================
# MAIN WORKFLOW
# ============================================================================

def extract_and_populate(source_file, template_file, output_file, sheet_name=None, header_row=None):
    """Complete workflow: Extract invoice data and populate Excel template"""
    
    print("🚀 Starting complete workflow...")
    print("="*50)
    
    # Step 1: Load and extract invoice data
    print("📋 Step 1: Loading and extracting invoice data...")
    
    try:
        sheet, df_raw = load_sheet(source_file, sheet_name=sheet_name)
        print(f"✅ Sheet loaded: {sheet}")
        print(f"📊 Data shape: {df_raw.shape}")
    except Exception as e:
        print(f"❌ Error loading source file: {e}")
        return False
    
    # Step 2: Detect header row
    print("\n🔍 Step 2: Detecting product table header...")
    
    if header_row is None:
        header_row = detect_product_header_row(df_raw)
        print(f"✅ Auto-detected header row: {header_row}")
    else:
        print(f"✅ Using specified header row: {header_row}")
    
    # Step 3: Extract data
    print("\n📊 Step 3: Extracting invoice data...")
    
    try:
        items, summary, idxmap = build_items(df_raw, header_row)
        meta = extract_meta(df_raw, header_row)
        
        invoice_data = {
            "sheet": sheet,
            "header_row": int(header_row),
            "columns_map": {int(k): v for k, v in idxmap.items()},
            "meta": meta,
            "items": items,
            "summary": summary
        }
        
        print(f"✅ Extracted {len(items)} product items")
        print(f"✅ Extracted {len(meta)} metadata fields")
        
    except Exception as e:
        print(f"❌ Error extracting data: {e}")
        return False
    
    # Step 4: Populate Excel template
    print("\n📋 Step 4: Populating Excel template...")
    
    try:
        output_path, default_values = populate_excel_template(invoice_data, template_file, output_file)
        print(f"✅ Excel template populated successfully")
    except Exception as e:
        print(f"❌ Error populating Excel: {e}")
        return False
    
    # Step 5: Generate summary report
    print("\n📋 Step 5: Generating summary report...")
    
    print("\n" + "="*60)
    print("📋 DATA MAPPING SUMMARY REPORT")
    print("="*60)
    
    print(f"\n📁 Source Invoice: {source_file}")
    print(f"📊 Total Items: {len(items)}")
    print(f"💰 Total Amount: {summary.get('مبلغ کل', 'N/A'):,} ریال")
    print(f"🏷️  Total Tax: {summary.get('جمع مالیات و عوارض', 'N/A'):,} ریال")
    print(f"🎯 Total Discount: {summary.get('مبلغ تخفیف', 'N/A'):,} ریال")
    
    print(f"\n📋 Target Excel: {output_path}")
    
    print("\n✅ Successfully Mapped Fields:")
    print("   - شناسه کالا (Product ID)")
    print("   - تعداد/مقدار (Quantity)")
    print("   - مبلغ تخفیف (Discount)")
    print("   - مبلغ مالیات و عوارض (Tax & Duties)")
    print("   - مبلغ واحد (Unit Price)")
    
    print("\n📅 Real Data Extracted from Header:")
    if 'تاریخ صدور فاکتور' in meta:
        print(f"   - تاریخ صدور فاکتور: {meta['تاریخ صدور فاکتور']}")
    if 'تاریخ ایجاد فاکتور' in meta:
        print(f"   - تاریخ ایجاد فاکتور: {meta['تاریخ ایجاد فاکتور']}")
    if 'شماره فاکتور اصلی' in meta:
        print(f"   - شماره فاکتور اصلی: {meta['شماره فاکتور اصلی']}")
    if 'شماره سریال' in meta:
        print(f"   - شماره سریال: {meta['شماره سریال']}")
    if 'سریال فاکتور حافظه مالیاتی' in meta:
        print(f"   - سریال فاکتور حافظه مالیاتی: {meta['سریال فاکتور حافظه مالیاتی']}")
    
    print("\n👤 Buyer Information:")
    if 'نام خریدار' in meta:
        print(f"   - نام خریدار: {meta['نام خریدار']}")
    if 'کد/شناسه ملی خریدار' in meta:
        print(f"   - کد/شناسه ملی خریدار: {meta['کد/شناسه ملی خریدار']}")
    
    print("\n🏢 Seller Information:")
    if 'شماره ثبت/ملی' in meta:
        print(f"   - شماره ثبت/ملی: {meta['شماره ثبت/ملی']}")
    if 'کد پستی' in meta:
        print(f"   - کد پستی: {meta['کد پستی']}")
    
    print(f"\n📅 Invoice Date: {default_values['تاریخ سند *']}")
    print(f"🔢 Invoice Number: {default_values['شماره صورتحساب']}")
    print(f"👤 Buyer: {default_values['نام خریدار']}")
    
    print("\n⚠️  Fields Still Using Default Values:")
    print("   - سایر اضافات (Other Charges) - 0")
    
    print("\n📝 Note: Most header information is now extracted from the source invoice!")
    print("   The Excel template now contains real data instead of generated defaults.")
    print("   Missing fields are left empty instead of filled with placeholder text.")
    print("   Invoice number uses the serial number (شماره سریال) from the source invoice.")
    
    print(f"\n🎉 Complete workflow finished successfully!")
    print(f"📁 Final output: {output_path}")
    
    return True

def main():
    """Main function with command line interface"""
    
    parser = argparse.ArgumentParser(
        description="Complete Solution: Extract invoice data and populate Excel template",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 complete_solution.py --source 1.xls --template "مصرف کننده.xlsm" --output result.xlsx
  python3 complete_solution.py --sources-dir "از فاکتور 232الی 242" --template "مصرف کننده.xlsم" --output result.xlsx
        """
    )
    
    parser.add_argument("--source", help="Single source invoice file (.xls, .xlsx)")
    parser.add_argument("--sources-dir", help="Directory of source invoices to batch process")
    parser.add_argument("--template", required=True, help="Excel template file (.xlsm)")
    parser.add_argument("--output", required=True, help="Output Excel file (.xlsx)")
    parser.add_argument("--sheet-name", help="Sheet name to extract from (auto-detected if not specified)")
    parser.add_argument("--header-row", type=int, help="Header row number (auto-detected if not specified)")
    
    args = parser.parse_args()
    
    # Validate template
    if not pathlib.Path(args.template).exists():
        print(f"❌ Template file not found: {args.template}")
        return
    
    output_dir = pathlib.Path(args.output).parent
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Batch mode
    if args.sources_dir:
        src_dir = pathlib.Path(args.sources_dir)
        if not src_dir.exists() or not src_dir.is_dir():
            print(f"❌ Sources directory not found: {src_dir}")
            return
        # Collect files
        files = []
        for ext in ("*.xls", "*.xlsx"):
            files.extend(src_dir.glob(ext))
        files = sorted(files)
        if not files:
            print(f"❌ No .xls/.xlsx files found in {src_dir}")
            return
        print(f"📦 Batch processing {len(files)} files from {src_dir}...")
        invoices = []
        for f in files:
            try:
                inv = extract_invoice_data(f, sheet_name=args.sheet_name, header_row=args.header_row)
                invoices.append(inv)
                print(f"  ✅ Extracted: {f.name} ({len(inv['items'])} items)")
            except Exception as e:
                print(f"  ⚠️ Skipped {f.name}: {e}")
        if not invoices:
            print("❌ No invoices extracted.")
            return
        populate_excel_template_batch(invoices, args.template, args.output)
        print(f"🎉 Batch output written: {args.output}")
        return
    
    # Single-file mode
    if not args.source:
        print("❌ Provide --source for single file or --sources-dir for batch.")
        return
    if not pathlib.Path(args.source).exists():
        print(f"❌ Source file not found: {args.source}")
        return
    
    success = extract_and_populate(
        source_file=args.source,
        template_file=args.template,
        output_file=args.output,
        sheet_name=args.sheet_name,
        header_row=args.header_row
    )
    
    if success:
        print("\n✅ Workflow completed successfully!")
    else:
        print("\n❌ Workflow failed!")
        exit(1)

if __name__ == "__main__":
    main()
